"""
Common utility functions for check-in analysis.
"""

from pathlib import Path

import openpyxl


def parse_athletes_roster(filepath: str) -> list[dict]:
    """Parse athletes from xlsx into normalized dicts.
    
    Returns list of athlete dicts with: id, firstName, lastName, age, gender,
    events, bibNumber, contactName, contactPhone, contactWeChat, contactEmail.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet_names = ["All", "Athletes 2025-05-22"]
    ws = None
    for name in sheet_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        raise ValueError(f"No valid sheet found. Available: {wb.sheetnames}")
    
    rows = list(ws.iter_rows(values_only=True))
    athletes = []
    
    for row in rows[1:]:
        if not row[3] or not row[4]:
            continue
        
        events = []
        if row[5]:
            events.append("4×100m Relay")
        if row[6]:
            events.append("100m")
        if row[7]:
            events.append("400m")
        if row[8]:
            events.append("1600m")
        
        athlete = {
            "id": str(int(row[2])),
            "firstName": str(row[3]).strip(),
            "lastName": str(row[4]).strip(),
            "age": int(row[0]),
            "gender": str(row[1]).strip(),
            "events": events,
            "bibNumber": int(row[2]),
        }
        athletes.append(athlete)
    
    wb.close()
    return athletes
