"""
EmP Check-In — Merge Check-In Logs

Collects JSON check-in logs exported from multiple devices via AirDrop,
replays events in timestamp order, and produces a final CSV report.

Usage:
    uv run python scripts/merge_checkins.py --logs-dir <path_to_logs> --roster <athletes.xlsx or volunteers.xlsx> --output <output.csv>

Log format (exported from each device):
[
  {
    "id": "302",
    "action": "checkIn" | "undo",
    "operator": "Haifeng",
    "device": "Haifeng's iPhone",
    "timestamp": "2026-06-14T08:32:15Z"
  },
  ...
]
"""

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl


def load_logs(logs_dir: Path) -> list[dict]:
    """Load and combine all JSON log files from a directory."""
    all_events = []
    log_files = list(logs_dir.glob("*.json"))
    
    if not log_files:
        print(f"ERROR: No .json files found in {logs_dir}")
        sys.exit(1)
    
    print(f"📂 Found {len(log_files)} log file(s):")
    for f in log_files:
        with open(f) as fh:
            events = json.load(fh)
        print(f"   {f.name}: {len(events)} events")
        all_events.extend(events)
    
    # Sort by timestamp
    all_events.sort(key=lambda e: e["timestamp"])
    return all_events


def replay_events(events: list[dict]) -> dict[str, dict]:
    """
    Replay events in timestamp order. Last action wins.
    Returns dict of id -> final check-in state.
    """
    states: dict[str, dict] = {}
    
    for event in events:
        person_id = event["personId"]
        action = event["action"]
        
        if action == "checkIn":
            states[person_id] = {
                "checkedIn": True,
                "checkedInAt": event["timestamp"],
                "checkedInBy": event["operator"],
                "device": event["device"],
            }
        elif action == "undo":
            states[person_id] = {
                "checkedIn": False,
                "undoneAt": event["timestamp"],
                "undoneBy": event["operator"],
                "device": event["device"],
            }
    
    return states


def merge_athletes(roster_path: str, states: dict, output_path: str):
    """Merge check-in states with athlete roster and output CSV."""
    wb = openpyxl.load_workbook(roster_path, read_only=True)
    ws = wb["Athletes 2025-05-22"]
    rows = list(ws.iter_rows(values_only=True))
    
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        # Header
        writer.writerow([
            "Bib #", "First Name", "Last Name", "Age", "Gender",
            "Events", "Checked In", "Checked In At", "Checked In By", "Device",
        ])
        
        checked_in_count = 0
        total = 0
        
        for row in rows[1:]:
            if not row[3] or not row[4]:
                continue
            
            total += 1
            bib = str(int(row[2]))
            first_name = str(row[3]).strip()
            last_name = str(row[4]).strip()
            age = int(row[0])
            gender = str(row[1]).strip()
            
            events = []
            if row[5]: events.append("4×100m")
            if row[6]: events.append("100m")
            if row[7]: events.append("400m")
            if row[8]: events.append("1600m")
            
            state = states.get(bib, {})
            checked_in = state.get("checkedIn", False)
            if checked_in:
                checked_in_count += 1
            
            writer.writerow([
                bib,
                first_name,
                last_name,
                age,
                gender,
                ", ".join(events) if events else "",
                "Yes" if checked_in else "No",
                state.get("checkedInAt", ""),
                state.get("checkedInBy", ""),
                state.get("device", ""),
            ])
    
    wb.close()
    print(f"\n✅ Athletes report: {checked_in_count}/{total} checked in")
    print(f"   Written to: {output_path}")


def merge_volunteers(roster_path: str, states: dict, output_path: str):
    """Merge check-in states with volunteer roster and output CSV."""
    wb = openpyxl.load_workbook(roster_path, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "ID", "Name", "Role", "Checked In", "Checked In At",
            "Checked In By", "Device",
        ])
        
        checked_in_count = 0
        total = 0
        vol_id = 1
        
        for row in rows[1:]:
            if len(row) < 3 or not row[2]:
                continue
            
            name = str(row[2]).strip()
            role = str(row[0]).strip() if row[0] else ""
            email = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            phone = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            
            # Skip entries that have no role, no email, and no phone (same as prepare_roster.py)
            if not role and not email and not phone:
                continue
            
            total += 1
            vid = str(vol_id)
            
            state = states.get(vid, {})
            checked_in = state.get("checkedIn", False)
            if checked_in:
                checked_in_count += 1
            
            writer.writerow([
                vid,
                name,
                role,
                "Yes" if checked_in else "No",
                state.get("checkedInAt", ""),
                state.get("checkedInBy", ""),
                state.get("device", ""),
            ])
            vol_id += 1
    
    wb.close()
    print(f"\n✅ Volunteers report: {checked_in_count}/{total} checked in")
    print(f"   Written to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="EmP Check-In Log Merger")
    parser.add_argument("--logs-dir", required=True, help="Directory containing exported JSON logs")
    parser.add_argument("--roster", required=True, help="Original roster xlsx file")
    parser.add_argument("--output", required=True, help="Output CSV file path")
    parser.add_argument("--type", choices=["athletes", "volunteers"], required=True,
                        help="Type of roster to merge")
    args = parser.parse_args()
    
    logs_dir = Path(args.logs_dir)
    if not logs_dir.exists():
        print(f"ERROR: {logs_dir} does not exist")
        sys.exit(1)
    
    # Load and replay events
    print("📥 Loading check-in logs...")
    events = load_logs(logs_dir)
    print(f"\n🔄 Replaying {len(events)} events...")
    states = replay_events(events)
    
    checked_in = sum(1 for s in states.values() if s.get("checkedIn"))
    print(f"   Final state: {checked_in} checked in, {len(states) - checked_in} undone")
    
    # Ensure output directory exists
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    
    # Merge with roster
    if args.type == "athletes":
        merge_athletes(args.roster, states, args.output)
    else:
        merge_volunteers(args.roster, states, args.output)


if __name__ == "__main__":
    main()
