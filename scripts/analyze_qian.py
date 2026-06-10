#!/usr/bin/env python3
"""
EmP Qian (Volunteer) Check-In Data Analyzer

Reads exported check-in JSON files, analyzes volunteer check-ins,
and generates HTML report.

Usage:
    source .venv/bin/activate
    python scripts/analyze_qian.py --input-dir data/2026-06-07/qian/exports
"""

import argparse
import json
import os
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

PST = timezone(timedelta(hours=-7))
EVENT_START = datetime(2026, 6, 7, 7, 0, 0, tzinfo=PST)  # Volunteers arrive earlier


def parse_volunteers(xlsx_path: str) -> dict:
    """Parse volunteer roster from Excel file, mapping to groups by sheet name."""
    wb = openpyxl.load_workbook(xlsx_path)
    
    # First, build name -> group mapping from group sheets
    name_to_group = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == '总表0513':
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        
        # Find first/last name columns
        first_idx = None
        last_idx = None
        for i, h in enumerate(headers):
            if h and 'First' in str(h):
                first_idx = i
            if h and ('Last' in str(h) or 'last' in str(h)):
                last_idx = i
        
        if first_idx is not None and last_idx is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                first = row[first_idx] or ''
                last = row[last_idx] or ''
                if first or last:
                    full_name = f'{first} {last}'.strip()
                    if full_name and full_name not in name_to_group:
                        name_to_group[full_name] = sheet_name
    
    # Now parse main roster and assign groups
    ws = wb['总表0513']
    headers = [cell.value for cell in ws[1]]
    
    vol_no_idx = headers.index('Volunteer No.')
    who_idx = headers.index('Who')
    task_idx = headers.index('Task')
    
    volunteers = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vol_no = row[vol_no_idx]
        if vol_no:
            name = row[who_idx] or ''
            task = row[task_idx] or ''
            group = name_to_group.get(name, None)  # None if not found
            
            volunteers[str(vol_no)] = {
                'id': str(vol_no),
                'name': name,
                'task': task,
                'group': group,
            }
    return volunteers


def load_checkin_files(input_dir: str) -> dict:
    """Load all check-in JSON files from directory."""
    all_events = []
    files = []
    
    for fname in sorted(os.listdir(input_dir)):
        if not fname.endswith('.json'):
            continue
        fpath = os.path.join(input_dir, fname)
        with open(fpath) as f:
            events = json.load(f)
        
        all_events.extend(events)
        files.append({
            'filename': fname,
            'event_count': len(events),
        })
    
    return {'all_events': all_events, 'files': files}


def analyze_events(events: list, volunteers: dict) -> dict:
    """Analyze check-in events."""
    # Filter by time and action
    valid_checkins = []
    
    for e in events:
        if e.get('action') != 'checkIn':
            continue
        try:
            ts = datetime.fromisoformat(e['timestamp'].replace('Z', '+00:00'))
            if ts.astimezone(PST) >= EVENT_START:
                valid_checkins.append(e)
        except:
            pass
    
    # Unique checked-in volunteers
    checked_in_ids = set(e['personId'] for e in valid_checkins)
    all_ids = set(volunteers.keys())
    not_checked_in_ids = all_ids - checked_in_ids
    
    # Time analysis
    timestamps = []
    for e in valid_checkins:
        try:
            ts = datetime.fromisoformat(e['timestamp'].replace('Z', '+00:00'))
            timestamps.append(ts)
        except:
            pass
    
    minute_counts = Counter()
    for ts in timestamps:
        ts_pst = ts.astimezone(PST)
        minute_counts[ts_pst.strftime("%m-%d %H:%M")] += 1
    
    return {
        'total_volunteers': len(volunteers),
        'checked_in': len(checked_in_ids),
        'not_checked_in': len(not_checked_in_ids),
        'checked_in_ids': sorted(checked_in_ids, key=lambda x: int(x)),
        'not_checked_in_ids': sorted(not_checked_in_ids, key=lambda x: int(x)),
        'minute_counts': dict(sorted(minute_counts.items())),
        'valid_checkins': len(valid_checkins),
    }


def build_roster_by_group(analysis: dict, volunteers: dict) -> dict:
    """Build roster data grouped by task group with checked/missing bib numbers.
    Ungrouped volunteers are added as '其他 Others' group."""
    checked_set = set(analysis['checked_in_ids'])
    by_group = {}
    ungrouped_checked = []
    ungrouped_missing = []
    
    for vid, vol in volunteers.items():
        group = vol.get('group')
        if group is None:  # Track ungrouped volunteers
            if vid in checked_set:
                ungrouped_checked.append(vid)
            else:
                ungrouped_missing.append(vid)
            continue
        if group not in by_group:
            by_group[group] = {'checked': [], 'missing': []}
        
        if vid in checked_set:
            by_group[group]['checked'].append(vid)
        else:
            by_group[group]['missing'].append(vid)
    
    # Sort bib numbers within each group
    for group in by_group:
        by_group[group]['checked'] = sorted(by_group[group]['checked'], key=int)
        by_group[group]['missing'] = sorted(by_group[group]['missing'], key=int)
    
    # Sort groups by total count (descending)
    sorted_groups = dict(sorted(by_group.items(), key=lambda x: -(len(x[1]['checked']) + len(x[1]['missing']))))
    
    # Add ungrouped as "其他 Others" at the end
    if ungrouped_checked or ungrouped_missing:
        sorted_groups['其他 Others'] = {
            'checked': sorted(ungrouped_checked, key=int),
            'missing': sorted(ungrouped_missing, key=int)
        }
    
    return sorted_groups


def generate_html(data: dict, analysis: dict, volunteers: dict, output_path: str):
    """Generate HTML report."""
    total = analysis['total_volunteers']
    checked = analysis['checked_in']
    missing = analysis['not_checked_in']
    rate = (checked / total * 100) if total > 0 else 0
    
    now = datetime.now().strftime("%Y年%m月%d日 %H:%M")
    device_count = len(data['files'])
    
    # Prepare chart data
    minute_labels = json.dumps(list(analysis['minute_counts'].keys()))
    minute_values = json.dumps(list(analysis['minute_counts'].values()))
    
    # Build roster by group (ungrouped volunteers are in "其他 Others")
    roster_by_group = build_roster_by_group(analysis, volunteers)
    
    # Build group tabs and panels
    tabs = []
    panels = []
    for group, data_group in roster_by_group.items():
        checked_count = len(data_group['checked'])
        missing_count = len(data_group['missing'])
        total_group = checked_count + missing_count
        rate_group = (checked_count / total_group * 100) if total_group > 0 else 0
        
        # Sanitize group name for HTML id
        group_id = group.replace(' ', '_')
        
        tabs.append(f'<button class="group-tab" onclick="showGroup(\'{group_id}\')" id="tab-{group_id}">{group} ({checked_count}/{total_group}) {rate_group:.0f}%</button>')
        
        checked_bibs = " ".join(f'<span class="bib bib-checked">#{b}</span>' for b in data_group['checked'])
        missing_bibs = " ".join(f'<span class="bib bib-missing">#{b}</span>' for b in data_group['missing'])
        
        panels.append(f'''
        <div class="group-panel" id="panel-{group_id}">
            <div class="group-summary">
                <strong>{group}</strong> · {checked_count}/{total_group} 已签到 · {rate_group:.0f}%
            </div>
            <div class="bib-section">
                <div class="bib-label">✅ 已签到 Checked In ({checked_count})</div>
                <div class="bib-list">{checked_bibs if checked_bibs else '<span class="bib-empty">无 None</span>'}</div>
            </div>
            <div class="bib-section">
                <div class="bib-label">❌ 未签到 Missing ({missing_count})</div>
                <div class="bib-list">{missing_bibs if missing_bibs else '<span class="bib-empty">无 None</span>'}</div>
            </div>
        </div>''')
    
    # Build "All" panel with all volunteers
    all_checked_bibs = " ".join(f'<span class="bib bib-checked">#{b}</span>' for b in analysis['checked_in_ids'])
    all_missing_bibs = " ".join(f'<span class="bib bib-missing">#{b}</span>' for b in analysis['not_checked_in_ids'])
    
    tabs_html = '\n        '.join(tabs)
    panels_html = '\n'.join(panels)
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>EmP 义工签到报告 Volunteer Check-In Report</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        :root {{
            --color-checked: #34c759;
            --color-missing: #8e8e93;
            --color-activity: #5ac8fa;
            --color-text: #1d1d1f;
            --color-muted: #86868b;
            --color-border: #e5e5ea;
            --color-bg: #f5f5f7;
        }}
        * {{ box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
               margin: 0; padding: 20px; background: var(--color-bg); color: var(--color-text); }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        h1 {{ color: var(--color-text); margin-bottom: 8px; }}
        h2 {{ color: #424245; margin-top: 32px; margin-bottom: 16px; 
             border-bottom: 2px solid var(--color-checked); padding-bottom: 8px; }}
        .subtitle {{ color: var(--color-muted); margin-bottom: 12px; font-size: 14px; }}
        .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
                      gap: 16px; margin-bottom: 24px; }}
        .stat-card {{ background: white; border-radius: 12px; padding: 20px;
                     box-shadow: 0 2px 8px rgba(0,0,0,0.08); text-align: center; }}
        .stat-value {{ font-size: 42px; font-weight: 700; color: var(--color-text); }}
        .stat-value.checked {{ color: var(--color-checked); }}
        .stat-value.missing {{ color: var(--color-missing); }}
        .stat-label {{ color: var(--color-muted); font-size: 13px; margin-top: 4px; }}
        .stat-highlight {{ background: linear-gradient(135deg, #34c759 0%, #30d158 100%); }}
        .stat-highlight .stat-value {{ color: white; }}
        .stat-highlight .stat-label {{ color: rgba(255,255,255,0.9); }}
        .chart-card {{ background: white; border-radius: 12px; padding: 20px;
                      box-shadow: 0 2px 8px rgba(0,0,0,0.08); margin-bottom: 24px; }}
        .chart-title {{ font-weight: 600; margin-bottom: 4px; }}
        .chart-subtitle {{ font-size: 12px; color: var(--color-muted); margin-bottom: 16px; }}
        .roster-card {{ background: white; border-radius: 12px; padding: 20px;
                       box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
        .bib-section {{ margin-bottom: 16px; }}
        .bib-label {{ font-size: 13px; font-weight: 600; margin-bottom: 8px; }}
        .bib-list {{ display: flex; flex-wrap: wrap; gap: 6px; }}
        .bib {{ display: inline-block; padding: 4px 8px; border-radius: 4px; font-size: 12px; font-family: monospace; }}
        .bib-checked {{ background: #e8f5e9; color: #2e7d32; }}
        .bib-missing {{ background: #fafafa; color: #757575; }}
        .bib-empty {{ color: var(--color-muted); font-style: italic; }}
        .nav-links {{ margin-bottom: 16px; display: flex; gap: 16px; }}
        .nav-links a {{ color: #007aff; text-decoration: none; font-size: 14px; }}
        .nav-links a:hover {{ text-decoration: underline; }}
        /* Group tabs */
        .group-tabs {{ display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 16px; }}
        .group-tab {{ padding: 8px 16px; border: 1px solid var(--color-border); border-radius: 20px;
                     background: white; cursor: pointer; font-size: 13px; transition: all 0.2s; }}
        .group-tab:hover {{ border-color: var(--color-checked); }}
        .group-tab.active {{ background: var(--color-checked); color: white; border-color: var(--color-checked); }}
        .group-panel {{ display: none; background: white; border-radius: 12px; padding: 20px;
                       box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
        .group-panel.active {{ display: block; }}
        .group-summary {{ font-size: 14px; color: var(--color-muted); margin-bottom: 16px; }}
    </style>
</head>
<body>
<div class="container">
    <nav class="nav-links">
        <a href="./">← 返回 Home</a>
    </nav>
    <h1>🙋 EmP 义工签到报告 Volunteer Check-In Report</h1>
    <p class="subtitle">生成时间 {now} · {device_count} 台设备 (iPhone)</p>
    
    <h2>📊 数据总览 At a Glance</h2>
    <div class="stats-grid">
        <div class="stat-card stat-highlight">
            <div class="stat-value">{rate:.0f}%</div>
            <div class="stat-label">签到率 Check-In Rate</div>
        </div>
        <div class="stat-card">
            <div class="stat-value checked">{checked}</div>
            <div class="stat-label">已签到 Checked In</div>
        </div>
        <div class="stat-card">
            <div class="stat-value missing">{missing}</div>
            <div class="stat-label">未签到 Missing</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{total}</div>
            <div class="stat-label">义工总数 Total</div>
        </div>
    </div>
    
    <h2>📈 签到动态 Check-In Activity</h2>
    <div class="chart-card">
        <div class="chart-title">💓 签到心跳图 Live Check-In Flow</div>
        <div class="chart-subtitle">每分钟签到人数</div>
        <canvas id="heartbeatChart"></canvas>
    </div>
    
    <h2>📋 按小组查看 Roster by Group</h2>
    <div class="group-tabs">
        <button class="group-tab active" onclick="showGroup('all')" id="tab-all">全部 All ({checked}/{total}) {rate:.0f}%</button>
        {tabs_html}
    </div>
    <div id="panel-all" class="group-panel active">
        <div class="group-summary">
            <strong>全部 All</strong> · {checked}/{total} 已签到 · {rate:.0f}%
        </div>
        <div class="bib-section">
            <div class="bib-label">✅ 已签到 Checked In ({checked})</div>
            <div class="bib-list">{all_checked_bibs}</div>
        </div>
        <div class="bib-section">
            <div class="bib-label">❌ 未签到 Missing ({missing})</div>
            <div class="bib-list">{all_missing_bibs}</div>
        </div>
    </div>
    {panels_html}
</div>

<script>
    new Chart(document.getElementById('heartbeatChart'), {{
        type: 'bar',
        data: {{
            labels: {minute_labels},
            datasets: [{{ label: '签到', data: {minute_values},
                backgroundColor: '#5ac8fa', borderRadius: 2 }}]
        }},
        options: {{
            responsive: true,
            plugins: {{ legend: {{ display: false }} }},
            scales: {{
                x: {{ title: {{ display: true, text: '时间 Time (PST)' }}, ticks: {{ maxRotation: 90, minRotation: 45 }} }},
                y: {{ beginAtZero: true, title: {{ display: true, text: '每分钟人数' }} }}
            }}
        }}
    }});

    function showGroup(group) {{
        document.querySelectorAll('.group-panel').forEach(p => p.classList.remove('active'));
        document.querySelectorAll('.group-tab').forEach(t => t.classList.remove('active'));
        document.getElementById('panel-' + group).classList.add('active');
        document.getElementById('tab-' + group).classList.add('active');
    }}
</script>
</body>
</html>'''
    
    with open(output_path, 'w') as f:
        f.write(html)
    print(f"   Written HTML report to {output_path}")


def main():
    parser = argparse.ArgumentParser(description="EmP Qian Check-In Analyzer")
    parser.add_argument("--input-dir", required=True, help="Directory with exported JSON")
    parser.add_argument("--data-dir", default="data", help="Directory with volunteers.xlsx")
    parser.add_argument("--output", default=None, help="Output HTML file path")
    args = parser.parse_args()
    
    # Load volunteer roster
    print("📋 Loading volunteer roster...")
    volunteers_file = Path(args.data_dir) / "volunteers.xlsx"
    if not volunteers_file.exists():
        print(f"ERROR: {volunteers_file} not found")
        return
    
    volunteers = parse_volunteers(str(volunteers_file))
    print(f"   Loaded {len(volunteers)} volunteers")
    
    # Load check-in files
    print("📂 Loading check-in exports...")
    data = load_checkin_files(args.input_dir)
    print(f"   Loaded {len(data['files'])} files with {len(data['all_events'])} events")
    
    # Analyze
    print("🔍 Analyzing...")
    analysis = analyze_events(data['all_events'], volunteers)
    
    print(f"\n📊 Summary:")
    print(f"   Checked in: {analysis['checked_in']} / {analysis['total_volunteers']}")
    print(f"   Missing: {analysis['not_checked_in']}")
    
    # Generate HTML report
    output_path = args.output or str(Path(args.input_dir).parent / "report.html")
    print(f"\n📝 Generating HTML report...")
    generate_html(data, analysis, volunteers, output_path)
    
    print(f"\n✅ Done! Open: file://{os.path.abspath(output_path)}")


if __name__ == "__main__":
    main()
