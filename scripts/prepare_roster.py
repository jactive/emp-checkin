"""
EmP Check-In — Roster Preprocessing Script

Reads athletes.xlsx and volunteers.xlsx, normalizes data into JSON,
encrypts with AES-GCM using a shared password, and outputs .enc files
ready to bundle into the iOS apps.

Also detects and reports potential duplicate registrations.

Usage:
    uv run python scripts/prepare_roster.py --password <shared_password>
"""

import argparse
import hashlib
import json
import os
import sys
from collections import Counter
from pathlib import Path

import openpyxl

# AES-GCM encryption using cryptography library
try:
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
except ImportError:
    print("ERROR: 'cryptography' package required. Run: uv add cryptography")
    sys.exit(1)


def derive_key(password: str) -> bytes:
    """Derive a 256-bit AES key from password using SHA-256.
    
    Note: Using SHA-256 for simplicity. The iOS app must use the same
    derivation method. For a local-only community event app with a shared
    password, this is sufficient.
    """
    return hashlib.sha256(password.encode("utf-8")).digest()


def encrypt_data(data: bytes, password: str) -> bytes:
    """Encrypt data with AES-256-GCM. Returns nonce (12 bytes) + ciphertext."""
    key = derive_key(password)
    nonce = os.urandom(12)
    aesgcm = AESGCM(key)
    ciphertext = aesgcm.encrypt(nonce, data, None)
    return nonce + ciphertext


def _clean_phone(value) -> str:
    """Clean phone number from Excel — strip .0 from float representation."""
    if not value:
        return ""
    s = str(value).strip()
    # Excel stores numbers as floats, e.g. 4255984859.0
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _clean_str(value) -> str:
    """Clean string value from Excel."""
    if not value:
        return ""
    return str(value).strip()


def parse_athletes(filepath: str) -> list[dict]:
    """Parse athletes from xlsx into normalized dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    # Try different sheet names for compatibility
    sheet_names = ["All", "Athletes 2025-05-22"]
    ws = None
    for name in sheet_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        raise ValueError(f"No valid sheet found. Available: {wb.sheetnames}")
    rows = list(ws.iter_rows(values_only=True))
    
    athletes = []
    for row in rows[1:]:
        # Skip empty rows
        if not row[3] or not row[4]:
            continue
        
        # Parse events
        events = []
        if row[5]:  # 4x100m (9+)
            events.append("4×100m Relay")
        if row[6]:  # 100m (15+)
            events.append("100m")
        if row[7]:  # 400m (15+)
            events.append("400m")
        if row[8]:  # 1600m (11+)
            events.append("1600m")
        
        athlete = {
            "id": str(int(row[2])),  # Bib number as string ID
            "firstName": str(row[3]).strip(),
            "lastName": str(row[4]).strip(),
            "age": int(row[0]),
            "gender": str(row[1]).strip(),
            "events": events,
            "bibNumber": int(row[2]),
            # Contact info (only visible in EmP 通)
            "contactName": str(row[10]).strip() if row[10] else "",
            "contactPhone": _clean_phone(row[11]),
            "contactWeChat": str(row[12]).strip() if row[12] else "",
            "contactEmail": str(row[13]).strip() if row[13] else "",
            "shippingPhone": _clean_phone(row[14]),
        }
        athletes.append(athlete)
    
    wb.close()
    return athletes


def _normalize_name(name: str) -> str:
    """Normalize name for matching: lowercase, strip spaces, remove punctuation."""
    if not name:
        return ""
    return name.lower().strip().replace(" ", "").replace("-", "").replace(".", "")


def _format_header(header: str) -> str:
    """Format header to Title Case (e.g., 'Last name' -> 'Last Name')."""
    if not header:
        return ""
    return " ".join(word.capitalize() for word in header.strip().split())


def parse_volunteers(filepath: str) -> list[dict]:
    """Parse volunteers from xlsx into normalized dicts.
    
    Expected format (from 总表0513 sheet):
    - Column A: Task (role)
    - Column B: Volunteer No.
    - Column C: Who (name)
    - Column D: Other roles
    - Column E: Email
    - Column F: Phone
    - Column G: WeChat
    
    Other sheets contain group assignments:
    - Row 0 = headers (Task, First Name, Last Name, other columns...)
    - Tab name = Group name
    - Columns marked "ignored" are skipped
    - Other columns displayed as "Header: Value"
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    
    # Try to find the main volunteer sheet
    main_sheet_names = ["总表0513", "Sheet1"]
    main_ws = None
    main_sheet_name = None
    for name in main_sheet_names:
        if name in wb.sheetnames:
            main_ws = wb[name]
            main_sheet_name = name
            break
    if main_ws is None:
        raise ValueError(f"No valid volunteer sheet found. Available: {wb.sheetnames}")
    
    rows = list(main_ws.iter_rows(values_only=True))
    
    # First pass: collect all entries by volunteer number from main sheet
    vol_data: dict[int, dict] = {}
    
    for row in rows[1:]:  # Skip header
        # Need at least volunteer number and name
        if len(row) < 3:
            continue
        
        vol_no = row[1]
        name = _clean_str(row[2])
        
        if not vol_no or not name:
            continue
        
        # Convert volunteer number to int
        try:
            vol_no_int = int(float(vol_no))
        except (ValueError, TypeError):
            continue
        
        task = _clean_str(row[0]) if len(row) > 0 else ""
        other_roles = _clean_str(row[3]) if len(row) > 3 else ""
        email = _clean_str(row[4]) if len(row) > 4 else ""
        phone = _clean_phone(row[5]) if len(row) > 5 else ""
        wechat = _clean_str(row[6]) if len(row) > 6 else ""
        
        if vol_no_int not in vol_data:
            # First occurrence - create new entry
            vol_data[vol_no_int] = {
                "id": str(vol_no_int),
                "volunteerNumber": vol_no_int,
                "name": name,
                "tasks": [task] if task else [],
                "otherRoles": other_roles,
                "email": email,
                "phone": phone,
                "wechat": wechat,
                "groupName": "",
                "groupInfo": [],  # List of "Header: Value" strings
            }
        else:
            # Duplicate - merge tasks
            if task and task not in vol_data[vol_no_int]["tasks"]:
                vol_data[vol_no_int]["tasks"].append(task)
            # Fill in missing contact info if available
            if not vol_data[vol_no_int]["email"] and email:
                vol_data[vol_no_int]["email"] = email
            if not vol_data[vol_no_int]["phone"] and phone:
                vol_data[vol_no_int]["phone"] = phone
            if not vol_data[vol_no_int]["wechat"] and wechat:
                vol_data[vol_no_int]["wechat"] = wechat
            if not vol_data[vol_no_int]["otherRoles"] and other_roles:
                vol_data[vol_no_int]["otherRoles"] = other_roles
    
    # Build name-to-volunteer-number lookup (normalized name -> vol_no)
    name_to_vol: dict[str, int] = {}
    for vol_no_int, v in vol_data.items():
        norm_name = _normalize_name(v["name"])
        if norm_name:
            name_to_vol[norm_name] = vol_no_int
    
    # Second pass: read other sheets to find group assignments
    print(f"   Scanning {len(wb.sheetnames) - 1} group sheets for assignments...")
    matches_found = 0
    
    for sheet_name in wb.sheetnames:
        if sheet_name == main_sheet_name:
            continue  # Skip main sheet
        
        group_name = sheet_name  # Tab name is the group name
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        if len(rows) < 2:
            continue  # Need at least header + 1 data row
        
        # Parse header row (row 0)
        header_row = rows[0]
        headers = [_clean_str(h) for h in header_row] if header_row else []
        
        # Find column indices for First Name and Last Name
        first_name_idx = None
        last_name_idx = None
        for idx, h in enumerate(headers):
            h_lower = h.lower().replace(" ", "")
            if h_lower == "firstname":
                first_name_idx = idx
            elif h_lower == "lastname":
                last_name_idx = idx
        
        if first_name_idx is None:
            continue  # Can't match without first name
        
        # Process data rows
        for row in rows[1:]:
            if len(row) <= first_name_idx:
                continue
            
            first_name = _clean_str(row[first_name_idx])
            if not first_name:
                continue
            
            last_name = ""
            if last_name_idx is not None and len(row) > last_name_idx:
                last_name = _clean_str(row[last_name_idx])
            
            # Construct full name and try to match
            full_name = f"{first_name} {last_name}".strip() if last_name else first_name
            norm_full = _normalize_name(full_name)
            
            if norm_full not in name_to_vol:
                continue
            
            vol_no = name_to_vol[norm_full]
            
            # Only set if not already set (first match wins)
            if vol_data[vol_no]["groupName"]:
                continue
            
            vol_data[vol_no]["groupName"] = group_name
            
            # Collect additional info from other columns
            group_info = []
            for idx, h in enumerate(headers):
                # Skip ignored, empty headers, first name, last name
                h_lower = h.lower().replace(" ", "")
                if not h or h_lower == "ignored" or h_lower == "firstname" or h_lower == "lastname":
                    continue
                
                # Get cell value
                if idx < len(row):
                    value = _clean_str(row[idx])
                    if value:
                        formatted_header = _format_header(h)
                        group_info.append(f"{formatted_header}: {value}")
            
            vol_data[vol_no]["groupInfo"] = group_info
            matches_found += 1
    
    print(f"   Found group assignments for {matches_found} volunteers")
    
    # Convert to final format
    volunteers = []
    for vol_no_int in sorted(vol_data.keys()):
        v = vol_data[vol_no_int]
        volunteers.append({
            "id": v["id"],
            "volunteerNumber": v["volunteerNumber"],
            "name": v["name"],
            "role": " | ".join(v["tasks"]) if v["tasks"] else "",
            "otherRoles": v["otherRoles"],
            "groupName": v["groupName"],
            "groupInfo": v["groupInfo"],  # List of "Header: Value" strings
            "email": v["email"],
            "phone": v["phone"],
            "wechat": v["wechat"],
        })
    
    wb.close()
    return volunteers


def detect_athlete_duplicates(athletes: list[dict]) -> list[str]:
    """Detect potential duplicate athlete registrations."""
    warnings = []
    
    # Group by (firstName, lastName, age, gender)
    groups: dict[tuple, list[dict]] = {}
    for a in athletes:
        key = (
            a["firstName"].lower(),
            a["lastName"].lower(),
            a["age"],
            a["gender"],
        )
        groups.setdefault(key, []).append(a)
    
    for key, group in groups.items():
        if len(group) > 1:
            # Check if same contact (likely duplicate registration)
            contacts = set()
            for a in group:
                contact_key = (
                    a["contactEmail"].lower(),
                    a["shippingPhone"],
                )
                contacts.add(contact_key)
            
            if len(contacts) == 1:
                msg = (
                    f"⚠️  LIKELY DUPLICATE: {group[0]['firstName']} {group[0]['lastName']}, "
                    f"age {group[0]['age']} {group[0]['gender']} — "
                    f"{len(group)} registrations with same contact. "
                    f"Bibs: {[a['bibNumber'] for a in group]}"
                )
            else:
                msg = (
                    f"ℹ️  SAME NAME: {group[0]['firstName']} {group[0]['lastName']}, "
                    f"age {group[0]['age']} {group[0]['gender']} — "
                    f"{len(group)} entries with DIFFERENT contacts (likely siblings/twins). "
                    f"Bibs: {[a['bibNumber'] for a in group]}"
                )
            warnings.append(msg)
    
    return warnings


def main():
    parser = argparse.ArgumentParser(description="EmP Check-In Roster Preprocessor")
    parser.add_argument("--password-dao", required=True, help="Password for EmP 到 (athlete app)")
    parser.add_argument("--password-qian", required=True, help="Password for EmP 签 (volunteer app)")
    parser.add_argument("--password-tong", required=True, help="Password for EmP 通 (admin app)")
    parser.add_argument("--data-dir", default="data", help="Directory containing xlsx files")
    parser.add_argument("--output-dir", default="app_bundle", help="Output directory for encrypted files")
    args = parser.parse_args()
    
    data_dir = Path(args.data_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(exist_ok=True)
    
    # Parse athletes
    print("📋 Parsing athletes...")
    athletes_file = data_dir / "athletes.xlsx"
    if not athletes_file.exists():
        print(f"ERROR: {athletes_file} not found")
        sys.exit(1)
    athletes = parse_athletes(str(athletes_file))
    print(f"   Found {len(athletes)} athletes")
    
    # Validate bib uniqueness
    bibs = [a["bibNumber"] for a in athletes]
    seen = {}
    for a in athletes:
        bib = a["bibNumber"]
        if bib in seen:
            print(f"   ❌ ERROR: Duplicate bib #{bib}: {seen[bib]} and {a['firstName']} {a['lastName']}")
            sys.exit(1)
        seen[bib] = f"{a['firstName']} {a['lastName']}"
    print(f"   ✓ All {len(bibs)} bibs are unique")
    
    # Parse volunteers
    print("📋 Parsing volunteers...")
    volunteers_file = data_dir / "volunteers.xlsx"
    if not volunteers_file.exists():
        print(f"ERROR: {volunteers_file} not found")
        sys.exit(1)
    volunteers = parse_volunteers(str(volunteers_file))
    print(f"   Found {len(volunteers)} volunteers")
    
    # Detect duplicates
    print("\n🔍 Checking for duplicates...")
    warnings = detect_athlete_duplicates(athletes)
    if warnings:
        print(f"   Found {len(warnings)} potential issues:\n")
        for w in warnings:
            print(f"   {w}")
    else:
        print("   No duplicates detected ✓")
    
    # Encrypt and write athletes (for EmP 到)
    print("\n🔐 Encrypting athletes roster (EmP 到)...")
    athletes_json = json.dumps(athletes, ensure_ascii=False, indent=None).encode("utf-8")
    athletes_enc = encrypt_data(athletes_json, args.password_dao)
    athletes_out = output_dir / "athletes_dao.enc"
    athletes_out.write_bytes(athletes_enc)
    print(f"   Written to {athletes_out} ({len(athletes_enc)} bytes)")
    
    # Encrypt and write volunteers (for EmP 签)
    print("🔐 Encrypting volunteers roster (EmP 签)...")
    volunteers_json = json.dumps(volunteers, ensure_ascii=False, indent=None).encode("utf-8")
    volunteers_enc = encrypt_data(volunteers_json, args.password_qian)
    volunteers_out = output_dir / "volunteers_qian.enc"
    volunteers_out.write_bytes(volunteers_enc)
    print(f"   Written to {volunteers_out} ({len(volunteers_enc)} bytes)")
    
    # Encrypt both for EmP 通 (admin sees everything)
    print("🔐 Encrypting both rosters (EmP 通)...")
    athletes_tong_enc = encrypt_data(athletes_json, args.password_tong)
    volunteers_tong_enc = encrypt_data(volunteers_json, args.password_tong)
    (output_dir / "athletes_tong.enc").write_bytes(athletes_tong_enc)
    (output_dir / "volunteers_tong.enc").write_bytes(volunteers_tong_enc)
    print(f"   Written athletes_tong.enc and volunteers_tong.enc")
    
    # Generate age metadata for EmP 到 (unencrypted, just the list of ages)
    print("📊 Generating age metadata...")
    available_ages = sorted(set(a["age"] for a in athletes))
    ages_json = json.dumps(available_ages).encode("utf-8")
    ages_out = output_dir / "ages_dao.json"
    ages_out.write_bytes(ages_json)
    print(f"   Available ages: {available_ages}")
    print(f"   Written to {ages_out}")
    
    # Summary
    print("\n✅ Done!")
    print(f"   Athletes:   {len(athletes)} records")
    print(f"   Volunteers: {len(volunteers)} records")
    print(f"\n   EmP 到 password: {args.password_dao}")
    print(f"   EmP 签 password: {args.password_qian}")
    print(f"   EmP 通 password: {args.password_tong}")


if __name__ == "__main__":
    main()
